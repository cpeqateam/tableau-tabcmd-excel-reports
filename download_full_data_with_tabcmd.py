import subprocess

'''Tabcmd'in yüklü olduğu klasör'''
tabcmd_path = "C:\\Users\\75575650\\AppData\\Local\\Programs\\Python\\Python312\\Scripts\\tabcmd.exe"

'''Tableau server bilgileri'''
tableau_server_url = "http://10.116.125.12:9090"
tableau_username = "75575650"
tableau_password = "75575650"

sheets=["/Wi-FiMercekBolgelereGonderilenListeler/Mesh_Map_CSV","/Wi-FiMercekBolgelereGonderilenListeler/Upsell_DSL_Bar_CSV","/Wi-FiMercekBolgelereGonderilenListeler/Upsell_FIBER_DSL_Bar2_1","/Wi-FiMercekBolgelereGonderilenListeler/IP_Change_MAP2_1"]
sheets_png=["/Wi-FiMercekBolgelereGonderilenListeler/MESH","/Wi-FiMercekBolgelereGonderilenListeler/UPSELL_DSL","/Wi-FiMercekBolgelereGonderilenListeler/UPSELL_DSL_FIBER","/Wi-FiMercekBolgelereGonderilenListeler/IP_CHANGE"]
outputs = ["MESH","UPSELL_DSL","UPSELL_DSL_FIBER","IP_CHANGE"]

'''Tableauya Login Olma Komutu ve Login Olma İşlemi'''
login = f"\"{tabcmd_path}\" login -s {tableau_server_url} -u {tableau_username} --password {tableau_password}"
subprocess.run(login, shell=True)

for i, sheet in enumerate(sheets):
    '''Export İşlem Komutları ve Export İşlemi'''

    export = f"\"{tabcmd_path}\" export \"{sheet}\" --csv -f \"{outputs[i]}.csv\" --username {tableau_username} --password {tableau_password}"
    subprocess.run(export, shell=True)

for i, sheet in enumerate(sheets_png):
    '''Export İşlem Komutları ve Export İşlemi'''

    export = f"\"{tabcmd_path}\" export \"{sheet}\" --png -f \"{outputs[i]}.png\" --username {tableau_username} --password {tableau_password}"
    subprocess.run(export, shell=True)

'''logout olma'''
logout = f"\"{tabcmd_path}\" logout"
subprocess.run(logout, shell=True)

import os
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, LineChart, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

# =========================
# 1) AYARLAR
# =========================
# Çalışacağın CSV dosyaları (indirilenler)
CSV_FILES = [
    "MESH.csv",
    "UPSELL_DSL.csv",
    "UPSELL_DSL_FIBER.csv",
    "IP_CHANGE.csv",
]

# Ortak kolon sıralaması (IP_CHANGE HARİÇ)
COMMON_ORDER = ["Bölge", "İl", "İlçe", "Müdürlük", "Altyapı", "Wi-Fi Versiyon", "Profil", "MAC", "SERVICENO"]

# IP_CHANGE için özel kolon sırası
IP_CHANGE_ORDER = ["Bölge", "İl", "İlçe", "Müdürlük", "Altyapı", "Wi-Fi Versiyon", "Profil", "IP Değişikliği", "MAC", "SERVICENO"]

# =========================
# 2) YARDIMCI FONKSİYONLAR
# =========================
def read_csv_safely(path: Path) -> pd.DataFrame:
    """
    CSV'yi güvenli şekilde oku:
    - Önce varsayılan ayarlar
    - Olmazsa tab-separated olarak dene
    - Türkçe karakterler için utf-8-sig denenir
    """
    try:
        return pd.read_csv(path, low_memory=False)
    except Exception:
        try:
            return pd.read_csv(path, sep="\t", low_memory=False)
        except Exception:
            return pd.read_csv(path, low_memory=False, encoding="utf-8-sig")

def autosize_columns(ws):
    """OpenPyXL worksheet: içerik uzunluklarına göre kolon genişliklerini ayarla."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for c in col_cells:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

def style_header_row(ws):
    """İlk satırı kalın, gri arka plan, merkezli ve ince kenarlıklarla biçimlendir."""
    header_fill = PatternFill("solid", fgColor="D3D3D3")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill
        cell.border = border
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

def write_df_to_region_sheets(df: pd.DataFrame, wb: Workbook):
    """
    DataFrame'i Bölge sütununa göre ayırıp her bölge için ayrı sayfa yazar.
    """
    for region, grp in df.groupby("Bölge", dropna=False):
        sheet_name = str(region) if pd.notna(region) else "Bilinmeyen Bölge"
        sheet_name = sheet_name[:31]  # Excel sheet isim limiti
        ws = wb.create_sheet(title=sheet_name)

        # Başlık
        ws.append(list(grp.columns))
        style_header_row(ws)

        # Satırlar
        for row in grp.itertuples(index=False):
            ws.append(list(row))

        autosize_columns(ws)

def make_summary_sheet(wb_path: Path):
    """
    Var olan workbook'u açar, tüm bölge sheet'lerini okuyup ÖZET sheet'ini oluşturur,
    özet tabloları ve grafikleri ekler.
    """
    wb = load_workbook(wb_path)

    # ÖZET sheet’i temizle/oluştur
    summary_name = "ÖZET"
    if summary_name in wb.sheetnames:
        del wb[summary_name]
    ws_sum = wb.create_sheet(summary_name)

    # Tüm bölge sheet’lerini tek DF’de topla
    # (Sadece ÖZET harici sayfaları oku)
    df_all_list = []
    for ws in wb.worksheets:
        if ws.title == summary_name:
            continue
        # Sheet’i DataFrame olarak oku
        data = ws.values
        headers = next(data)
        df_ws = pd.DataFrame(data, columns=headers)
        df_all_list.append(df_ws)

    if not df_all_list:
        wb.save(wb_path)
        return

    df_all = pd.concat(df_all_list, ignore_index=True)

    # Özet tablolar
    def _safe_count(df, by_col):
        if by_col not in df.columns:
            return pd.DataFrame(columns=[by_col, "Adet"])
        return df.groupby(by_col)["MAC"].count().reset_index().rename(columns={"MAC": "Adet"})

    wifi_sum = _safe_count(df_all, "Wi-Fi Versiyon")
    alty_sum = _safe_count(df_all, "Altyapı")
    prof_sum = _safe_count(df_all, "Profil")
    bolg_sum = _safe_count(df_all, "Bölge")

    def write_table(start_row, start_col, df_tbl: pd.DataFrame):
        # Başlıklar
        for j, col_name in enumerate(df_tbl.columns, start=start_col):
            ws_sum.cell(row=start_row, column=j, value=col_name)
        # Satırlar
        for i, row in enumerate(df_tbl.itertuples(index=False), start=start_row + 1):
            for j, v in enumerate(row, start=start_col):
                ws_sum.cell(row=i, column=j, value=v)

    import re
    from openpyxl.chart import PieChart, LineChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, get_column_letter
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
    from openpyxl.styles import Font

    # --- Özet tabloları yaz (grafikler bunlardan beslenecek) ---
    # Not: Tablo yazılacak ama birazdan tüm yazıları beyaza çekip görünmez yapacağız.
    write_table(1, 1,  wifi_sum)   # Wi-Fi tablo: A1
    write_table(1, 5,  alty_sum)   # Altyapı tablo: H1
    write_table(6, 1,  prof_sum)   # Profil tablo: A20
    write_table(6, 5,  bolg_sum)   # Bölge  tablo: A50

    # --- Profil etiketlerini hız değerine göre sırala (1 Gbps > 512 Mbps > 51 Mbps ...) ---
    def _mbps_value(label: str) -> float:
        s = str(label).strip().lower()
        m = re.search(r'[\d\.,]+', s)
        if not m: return float('inf')
        v = float(m.group(0).replace(',', '.'))
        if 'gbps' in s:  v *= 1000.0
        elif 'kbps' in s: v /= 1000.0
        return v

    if not prof_sum.empty and "Profil" in prof_sum.columns:
        prof_sum = prof_sum.copy()
        prof_sum["__o__"] = prof_sum["Profil"].apply(_mbps_value)
        prof_sum = prof_sum.sort_values("__o__").drop(columns="__o__")
        # tabloyu güncelle ki grafik doğru sırayı kullansın
        write_table(20, 1, prof_sum)

    # --- ÖZET sayfasındaki tabloları görünmez yap (beyaz font, no fill, no border) ---
    def make_range_white(ws, start_row, start_col, nrows, ncols):
        for r in range(start_row, start_row + nrows + 1):  # +1: başlık satırını da dahil
            for c in range(start_col, start_col + ncols):
                cell = ws.cell(row=r, column=c)
                cell.font = Font(color="FFFFFF", bold=False)  # beyaz, bold kapalı
                cell.fill = PatternFill(fill_type=None)       # dolgu yok
                cell.border = Border()                        # çerçeve yok

    if len(wifi_sum) > 0:  make_range_white(ws_sum, 1, 1,  len(wifi_sum), 2)   # A:B
    if len(alty_sum) > 0:  make_range_white(ws_sum, 1, 5,  len(alty_sum), 2)   # E:F
    if len(prof_sum) > 0:  make_range_white(ws_sum, 20, 1,  len(prof_sum), 2)   # A:B
    if len(bolg_sum) > 0:  make_range_white(ws_sum, 6, 5,  len(bolg_sum), 2)   # A:B

    # ÖZET sayfasında freeze yok, gridlines kalsın/kalmasın tercihine göre değiştirebilirsin
    # ws_sum.freeze_panes = None            # zaten hiç ayarlamadık
    # ws_sum.sheet_view.showGridLines = True  # istersen görülsün

    # --- Grafikleri verilen hücre aralıklarına tam oturtmak için yardımcı ---
    def _add_chart_span(ws, chart, tl_cell: str, br_cell: str):
        tl_col_letter, tl_row = coordinate_from_string(tl_cell)
        br_col_letter, br_row = coordinate_from_string(br_cell)
        tl_col = column_index_from_string(tl_col_letter)
        br_col = column_index_from_string(br_col_letter)
        anchor = TwoCellAnchor(
            _from=AnchorMarker(col=tl_col-1, row=tl_row-1),
            to=AnchorMarker(col=br_col-1,  row=br_row-1)
        )
        ws.add_chart(chart, anchor)

    from openpyxl.chart.layout import Layout, ManualLayout

    # Sheet gridline'ı kapat
    ws_sum.sheet_view.showGridLines = False

    # 1) Wi-Fi Pie — geniş alan + plot'u aşağı kaydır (A1–J26)
    if len(wifi_sum) > 0:
        pie_wifi = PieChart()
        # Wi-Fi tablo: A1:B...  (B=Adet, A=Kategori)
        pie_wifi.add_data(Reference(ws_sum, min_col=2, min_row=2, max_row=1+len(wifi_sum)), titles_from_data=False)
        pie_wifi.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=1+len(wifi_sum)))
        pie_wifi.title = "Wi-Fi Versiyon"
        pie_wifi.legend.position = "t"      # başlığın altında
        pie_wifi.legend.overlay  = False
        pie_wifi.dataLabels = DataLabelList()
        pie_wifi.dataLabels.showPercent   = True
        pie_wifi.dataLabels.showVal       = False
        pie_wifi.dataLabels.showCatName   = False
        pie_wifi.dataLabels.showSerName   = False   # Series1 kapalı
        pie_wifi.dataLabels.showLegendKey = False
        # Başlık ve legend çakışmasın diye çizim alanını aşağı kaydır
        pie_wifi.layout = Layout(manualLayout=ManualLayout(y=0.22, h=0.75))
        _add_chart_span(ws_sum, pie_wifi, "A1", "J26")

    # 2) Altyapı Pie — geniş alan + plot'u aşağı kaydır (L1–V26)
    if len(alty_sum) > 0:
        pie_alt = PieChart()
        # Altyapı tablo: E1:F... (F=Adet, E=Kategori)
        pie_alt.add_data(Reference(ws_sum, min_col=6, min_row=2, max_row=1+len(alty_sum)), titles_from_data=False)
        pie_alt.set_categories(Reference(ws_sum, min_col=5, min_row=2, max_row=1+len(alty_sum)))
        pie_alt.title = "Altyapı"
        pie_alt.legend.position = "t"
        pie_alt.legend.overlay  = False
        pie_alt.dataLabels = DataLabelList()
        pie_alt.dataLabels.showPercent   = True
        pie_alt.dataLabels.showVal       = False
        pie_alt.dataLabels.showCatName   = False
        pie_alt.dataLabels.showSerName   = False
        pie_alt.dataLabels.showLegendKey = False
        pie_alt.layout = Layout(manualLayout=ManualLayout(y=0.22, h=0.75))
        _add_chart_span(ws_sum, pie_alt, "L1", "V26")

    # 3) Profil Line — X: gerçek Profil isimleri, Y: Adet (A28–V48)
    #    Tablon: A6:B... => Kategoriler A7:A..., Değerler B7:B...
    if len(prof_sum) > 0:
        line = LineChart()
        line.smooth = False            # düz çizgi
        line.legend = None
        line.add_data(
            Reference(ws_sum, min_col=2, min_row=7, max_row=6+len(prof_sum)),
            titles_from_data=False
        )  # Y = Adet
        line.set_categories(
            Reference(ws_sum, min_col=1, min_row=7, max_row=6+len(prof_sum))
        )  # X = Profil etiketleri
        line.title = "Profil Dağılımı"
        line.x_axis.title = "Profil"
        line.y_axis.title = "Adet"
        line.x_axis.tickLblPos = "low"   # etiketler görünsün
        # (opsiyonel) küçük markerlar okunaklılık için
        try:
            from openpyxl.chart.marker import Marker
            if line.series:
                line.series[0].marker = Marker(symbol="circle"); line.series[0].marker.size = 5
        except Exception:
            pass
        _add_chart_span(ws_sum, line, "A28", "V48")

    # 4) Bölge Bar — X: gerçek Bölge isimleri, Y: Adet (A50–V78)
    #    Tablon: E6:F... => Kategoriler E7:E..., Değerler F7:F...
    if len(bolg_sum) > 0:
        bar = BarChart()
        bar.legend = None
        bar.add_data(
            Reference(ws_sum, min_col=6, min_row=7, max_row=6+len(bolg_sum)),
            titles_from_data=False
        )  # Y = Adet
        bar.set_categories(
            Reference(ws_sum, min_col=5, min_row=7, max_row=6+len(bolg_sum))
        )  # X = Bölge etiketleri
        bar.title = "Bölge Dağılımı"
        bar.x_axis.title = "Bölge"
        bar.y_axis.title = "Adet"
        bar.x_axis.tickLblPos = "low"
        _add_chart_span(ws_sum, bar, "A50", "V78")

    # ÖZET’te başlık/otamatik kolon genişliği uygulanmasın (görünürlüğü istemiyorsun)
    # style_header_row(ws_sum)   # çağırmıyoruz
    # autosize_columns(ws_sum)   # çağırmıyoruz

    wb.save(wb_path)

# =========================
# 3) ANA DÖNÜŞÜM
# =========================
def process_one_csv(csv_name: str):
    """
    Tek bir CSV'yi işler:
      - Kolonları istenen sıraya dizer (IP_CHANGE özel)
      - 'Bölge'ye göre alfabetik sıralar
      - Bölgelere göre sheet’lere ayırır
      - ÖZET sheet'i ve grafiklerini ekler
    Çıktı: aynı isimle .xlsx (örn MESH.xlsx)
    """
    csv_path = Path(csv_name)
    if not csv_path.exists():
        print(f"❌ {csv_name} bulunamadı, atlanıyor...")
        return

    # 1) CSV'yi oku
    df = read_csv_safely(csv_path)
    df = df.loc[:, ~df.columns.duplicated()]  # mükerrer kolonları temizle

    # 2) Kolon sırasını seç
    if csv_name.upper().startswith("IP_CHANGE"):
        desired = [c for c in IP_CHANGE_ORDER if c in df.columns]
    else:
        desired = [c for c in COMMON_ORDER if c in df.columns]

    if not desired:
        print(f"⚠️ {csv_name} içinde beklenen kolonlardan hiçbiri yok. Mevcut kolonlar: {list(df.columns)}")
        return

    df = df[desired]

    # 3) 'Bölge'ye göre alfabetik sırala (varsa)
    if "Bölge" in df.columns:
        df = df.sort_values(by="Bölge", kind="stable")

    # 4) Workbook oluştur ve bölgelere göre sayfalara yaz
    out_xlsx = csv_path.with_suffix(".xlsx")
    if out_xlsx.exists():
        out_xlsx.unlink()  # var ise sil (yeniden yaratacağız)

    wb = Workbook()
    # default "Sheet" sayfasını kaldır
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    write_df_to_region_sheets(df, wb)

    # Bölge bazlı sheet hiç oluşmadıysa (veri yoksa) en azından bir sayfa yazalım
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("Veri Yok")
        ws.append(list(df.columns))
        style_header_row(ws)

    wb.save(out_xlsx)

    # 5) ÖZET sayfasını ekle + grafikler
    make_summary_sheet(out_xlsx)

    print(f"✅ {csv_name} → {out_xlsx.name} (bölgelere ayrıldı + ÖZET oluşturuldu)")

# =========================
# 4) ÇALIŞTIR
# =========================
if __name__ == "__main__":
    for csv in CSV_FILES:
        process_one_csv(csv)
    print("\n🎉 Tamamdır! İlgili .xlsx dosyaları oluştu.")

